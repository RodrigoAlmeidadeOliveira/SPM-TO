"""
Script para popular o banco de dados com instrumentos, questões e tabelas de referência das planilhas SPM
"""
import openpyxl
from pathlib import Path
from app import db
from app.models import Instrumento, Dominio, Questao, TabelaReferencia, PlanoTemplateItem


def seed_database():
    """Popula o banco de dados com dados das planilhas"""
    doctos_path = Path(__file__).parent.parent / 'DOCTOS'

    # Definição dos instrumentos
    instrumentos_config = [
        {
            'codigo': 'SPM_5_12_CASA',
            'nome': 'SPM 5-12 anos - Casa',
            'idade_minima': 5,
            'idade_maxima': 12,
            'contexto': 'casa',
            'planilha': 'Planilha PEI SPM 5-12 - casa.xlsx'
        },
        {
            'codigo': 'SPM_5_12_ESCOLA',
            'nome': 'SPM 5-12 anos - Escola',
            'idade_minima': 5,
            'idade_maxima': 12,
            'contexto': 'escola',
            'planilha': 'Planilha PEI SPM 5-12 - escola.xlsx'
        },
        {
            'codigo': 'SPM_P_3_5_CASA',
            'nome': 'SPM-P 3-5 anos - Casa',
            'idade_minima': 3,
            'idade_maxima': 5,
            'contexto': 'casa',
            'planilha': 'Planilha PEI SPM-P 3-5 - casa.xlsx'
        },
        {
            'codigo': 'SPM_P_3_5_ESCOLA',
            'nome': 'SPM-P 3-5 anos - Escola',
            'idade_minima': 3,
            'idade_maxima': 5,
            'contexto': 'escola',
            'planilha': 'Planilha PEI SPM-P 3-5 - escola.xlsx'
        }
    ]

    # Definição dos domínios (ordem e características)
    dominios_config = {
        'SPM_5_12': [
            {'codigo': 'SOC', 'nome': 'Participação Social', 'ordem': 1, 'escala_invertida': False},
            {'codigo': 'VIS', 'nome': 'Visão', 'ordem': 2, 'escala_invertida': True},
            {'codigo': 'HEA', 'nome': 'Audição', 'ordem': 3, 'escala_invertida': True},
            {'codigo': 'TOU', 'nome': 'Tato', 'ordem': 4, 'escala_invertida': True},
            {'codigo': 'BOD', 'nome': 'Consciência Corporal', 'ordem': 5, 'escala_invertida': True},
            {'codigo': 'BAL', 'nome': 'Equilíbrio e Movimento', 'ordem': 6, 'escala_invertida': True},
            {'codigo': 'PLA', 'nome': 'Planejamento e Ideação', 'ordem': 7, 'escala_invertida': True}
        ],
        'SPM_P_3_5': [
            {'codigo': 'SOC', 'nome': 'Participação Social', 'ordem': 1, 'escala_invertida': False},
            {'codigo': 'VIS', 'nome': 'Visão', 'ordem': 2, 'escala_invertida': True},
            {'codigo': 'HEA', 'nome': 'Audição', 'ordem': 3, 'escala_invertida': True},
            {'codigo': 'OLF', 'nome': 'Olfato e Paladar', 'ordem': 4, 'escala_invertida': True},
            {'codigo': 'TOU', 'nome': 'Tato', 'ordem': 5, 'escala_invertida': True},
            {'codigo': 'BOD', 'nome': 'Consciência Corporal', 'ordem': 6, 'escala_invertida': True},
            {'codigo': 'BAL', 'nome': 'Equilíbrio e Movimento', 'ordem': 7, 'escala_invertida': True},
            {'codigo': 'PLA', 'nome': 'Planejamento e Ideação', 'ordem': 8, 'escala_invertida': True}
        ]
    }

    print('Iniciando população do banco de dados...')

    # Criar instrumentos
    for config in instrumentos_config:
        print(f"\nProcessando {config['nome']}...")

        instrumento = Instrumento.query.filter_by(codigo=config['codigo']).first()
        if instrumento:
            instrumento.nome = config['nome']
            instrumento.idade_minima = config['idade_minima']
            instrumento.idade_maxima = config['idade_maxima']
            instrumento.contexto = config['contexto']
            instrumento.instrucoes = 'Por favor, responda as perguntas deste formulário de acordo com a frequência.'
            print("  -> Instrumento já existia, atualizado.")
        else:
            instrumento = Instrumento(
                codigo=config['codigo'],
                nome=config['nome'],
                idade_minima=config['idade_minima'],
                idade_maxima=config['idade_maxima'],
                contexto=config['contexto'],
                instrucoes='Por favor, responda as perguntas deste formulário de acordo com a frequência.'
            )
            db.session.add(instrumento)
            db.session.flush()
        dominio_set = 'SPM_P_3_5' if 'SPM_P_3_5' in config['codigo'] else 'SPM_5_12'

        dominios_criados = []
        for dom_config in dominios_config[dominio_set]:
            dominio = Dominio.query.filter_by(instrumento_id=instrumento.id, codigo=dom_config['codigo']).first()
            if dominio:
                dominio.nome = dom_config['nome']
                dominio.ordem = dom_config['ordem']
                dominio.escala_invertida = dom_config['escala_invertida']
                print(f"  -> Domínio atualizado: {dom_config['nome']}")
            else:
                dominio = Dominio(
                    instrumento_id=instrumento.id,
                    codigo=dom_config['codigo'],
                    nome=dom_config['nome'],
                    ordem=dom_config['ordem'],
                    escala_invertida=dom_config['escala_invertida']
                )
                db.session.add(dominio)
                db.session.flush()
                print(f"  - Domínio criado: {dom_config['nome']}")
            dominios_criados.append(dominio)

        # Ler questões da planilha
        try:
            planilha_path = doctos_path / config['planilha']
            extrair_questoes_planilha(planilha_path, instrumento, dominios_criados)

            PlanoTemplateItem.query.filter_by(instrumento_id=instrumento.id).delete()
            db.session.flush()
            extrair_plano_planilha(planilha_path, instrumento, dominios_criados)

            # Extrair tabela de referência (se houver)
            extrair_tabela_referencia(planilha_path, instrumento)
        except Exception as e:
            print(f"  AVISO: Erro ao ler planilha: {e}")

    db.session.commit()
    print('\n✓ Banco de dados populado com sucesso!')


def extrair_questoes_planilha(planilha_path, instrumento, dominios_criados):
    """
    Extrai questões da planilha Excel

    Args:
        planilha_path: Caminho para a planilha
        instrumento: Instância de Instrumento
    """
    wb = openpyxl.load_workbook(planilha_path, data_only=True)
    ws = wb['Entrada de Dados']

    # Mapeamento de domínios
    dominios_por_nome = {d.nome.lower(): d for d in dominios_criados}

    # Variáveis de controle
    dominio_atual = None
    numero_questao = 1
    numero_questao_dominio = 1

    def extrair_texto(row):
        """Retorna o texto relevante da linha (considerando colunas mescladas)."""
        valores = [cell.value for cell in row]
        preferencia = [1, 2, 0, 3, 4, 5]

        for idx in preferencia:
            if idx < len(valores):
                valor = valores[idx]
                if isinstance(valor, str) and valor.strip():
                    return valor.strip()

        for valor in valores:
            if isinstance(valor, str) and valor.strip():
                return valor.strip()

        return None

    # Percorrer linhas da planilha (aproximadamente linhas 31 em diante)
    for row in ws.iter_rows(min_row=25, max_row=ws.max_row):
        texto = extrair_texto(row)
        if not texto:
            continue

        texto = texto.replace('\xa0', ' ').strip()
        texto_lower = texto.lower()

        if texto_lower.startswith('esta criança'):
            continue

        if texto_lower.startswith('score'):
            dominio_atual = None
            continue

        if texto[0].isdigit() and '.' in texto[:5]:
            _, possivel_nome = texto.split('.', 1)
            possivel_nome = possivel_nome.strip()
        else:
            possivel_nome = texto

        if possivel_nome.lower() in dominios_por_nome:
            dominio_atual = dominios_por_nome[possivel_nome.lower()]
            numero_questao_dominio = 1
            print(f"    Processando questões do domínio: {dominio_atual.nome}")
            continue

        if dominio_atual and len(texto) > 3:
            questao = Questao.query.join(Dominio).filter(
                Questao.dominio_id == dominio_atual.id,
                Questao.numero == numero_questao_dominio
            ).first()

            if questao:
                questao.texto = texto
            else:
                questao = Questao(
                    dominio_id=dominio_atual.id,
                    numero=numero_questao_dominio,
                    numero_global=numero_questao,
                    texto=texto
                )
                db.session.add(questao)

            numero_questao += 1
            numero_questao_dominio += 1

    print(f"    Total de questões extraídas: {numero_questao - 1}")


def extrair_plano_planilha(planilha_path, instrumento, dominios_criados):
    """
    Extrai itens de plano (PEI) da planilha Excel

    Args:
        planilha_path: Caminho para a planilha
        instrumento: Instância de Instrumento
        dominios_criados: Lista de domínios associados ao instrumento
    """
    wb = openpyxl.load_workbook(planilha_path)

    if 'Plano' not in wb.sheetnames:
        print("    AVISO: Planilha não possui aba 'Plano'")
        return

    ws = wb['Plano']

    dominios_por_nome = {d.nome.lower(): d for d in dominios_criados}
    dominio_atual = None
    ordem = 1

    for row in ws.iter_rows(min_row=3, values_only=True):
        texto = row[2] if len(row) >= 3 else None
        if not texto or not isinstance(texto, str):
            continue

        texto = texto.strip()
        if not texto:
            continue

        partes = texto.split('.', 1)

        if len(partes) == 2:
            resto = partes[1].strip().lower()
            if resto in dominios_por_nome:
                dominio_atual = dominios_por_nome[resto]
                continue

        if dominio_atual is None:
            continue

        item = PlanoTemplateItem(
            instrumento_id=instrumento.id,
            dominio_id=dominio_atual.id,
            ordem=ordem,
            texto=texto
        )
        db.session.add(item)
        ordem += 1

    print(f"    Itens de plano extraídos: {ordem - 1}")


def extrair_tabela_referencia(planilha_path, instrumento):
    """
    Extrai tabela de referência da planilha Excel

    Args:
        planilha_path: Caminho para a planilha
        instrumento: Instância de Instrumento
    """
    wb = openpyxl.load_workbook(planilha_path)

    # Verificar se tem aba TAB. REFERÊNCIA
    if 'TAB. REFERÊNCIA' not in wb.sheetnames:
        print(f"    AVISO: Planilha não tem aba 'TAB. REFERÊNCIA'")
        return

    ws = wb['TAB. REFERÊNCIA']

    # Encontrar cabeçalho (linha 3 geralmente tem: %TILE, T, SOC, VIS, HEA, TOU, BOD, BAL, PLA)
    # Mapeamento de colunas
    col_map = {}
    for idx, cell in enumerate(ws[3], 1):
        if cell.value:
            valor = str(cell.value).strip().upper()
            if valor == 'T':
                col_map['T'] = idx
            elif valor == '%TILE':
                col_map['PERCENTIL'] = idx
            elif valor in ['SOC', 'VIS', 'HEA', 'TOU', 'BOD', 'BAL', 'PLA', 'OLF']:
                col_map[valor] = idx

    print(f"    Extraindo tabela de referência...")
    print(f"    Colunas encontradas: {list(col_map.keys())}")

    # Percorrer linhas da tabela
    classificacao_atual = None
    count_refs = 0

    for row_idx in range(4, ws.max_row + 1):
        row = ws[row_idx]

        # Coluna B geralmente tem a classificação
        celula_b = row[1].value
        if celula_b:
            celula_b_str = str(celula_b).strip().upper()
            if 'DISFUNÇÃO DEFINITIVA' in celula_b_str or 'DISFUNCAO DEFINITIVA' in celula_b_str:
                classificacao_atual = 'DISFUNCAO_DEFINITIVA'
            elif 'PROVÁVEL DISFUNÇÃO' in celula_b_str or 'PROVAVEL DISFUNCAO' in celula_b_str:
                classificacao_atual = 'PROVAVEL_DISFUNCAO'
            elif 'TÍPICO' in celula_b_str or 'TIPICO' in celula_b_str:
                classificacao_atual = 'TIPICO'

        if not classificacao_atual or 'T' not in col_map:
            continue

        # Obter T-score
        t_score_cell = row[col_map['T'] - 1].value
        if not t_score_cell or not str(t_score_cell).strip().isdigit():
            continue

        t_score = int(t_score_cell)

        # Obter percentil
        percentil_min = None
        percentil_max = None
        if 'PERCENTIL' in col_map:
            percentil_cell = row[col_map['PERCENTIL'] - 1].value
            if percentil_cell:
                percentil_str = str(percentil_cell).strip()
                if percentil_str and percentil_str != '':
                    # Pode ser um número ou range (ex: ">99", "95-97")
                    if '>' in percentil_str:
                        percentil_min = 99
                        percentil_max = 100
                    elif '-' in percentil_str:
                        parts = percentil_str.split('-')
                        percentil_min = int(parts[0].strip())
                        percentil_max = int(parts[1].strip())
                    elif percentil_str.isdigit():
                        percentil_min = percentil_max = int(percentil_str)

        # Processar cada domínio
        for dominio_codigo in ['SOC', 'VIS', 'HEA', 'TOU', 'BOD', 'BAL', 'PLA', 'OLF']:
            if dominio_codigo not in col_map:
                continue

            cell_value = row[col_map[dominio_codigo] - 1].value
            if not cell_value:
                continue

            # Parse do valor (pode ser "37-40", "35-36", "33", etc.)
            value_str = str(cell_value).strip()
            if not value_str or value_str == '':
                continue

            escore_min = escore_max = None

            if '-' in value_str:
                parts = value_str.split('-')
                try:
                    escore_min = int(parts[0].strip())
                    escore_max = int(parts[1].strip())
                except:
                    continue
            elif value_str.isdigit():
                escore_min = escore_max = int(value_str)
            else:
                continue

            # Criar entrada na tabela de referência
            ref = TabelaReferencia(
                instrumento_id=instrumento.id,
                dominio_codigo=dominio_codigo,
                t_score=t_score,
                percentil_min=percentil_min,
                percentil_max=percentil_max,
                escore_min=escore_min,
                escore_max=escore_max,
                classificacao=classificacao_atual
            )
            db.session.add(ref)
            count_refs += 1

    print(f"    Total de referências extraídas: {count_refs}")
